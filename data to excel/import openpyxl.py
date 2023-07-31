import openpyxl
from datetime import datetime

def get_input():
    roll_number = input("Enter Roll Number: ")
    name = input("Enter Name: ")
    age = input("Enter Age: ")
    email = input("Enter Email: ")
    return roll_number, name, age, email

def validate_roll_number(roll_number):
    if not roll_number.isnumeric():
        print("Invalid Roll Number. Roll Number should be a numeric value.")
        return False
    return True

def validate_age(age):
    if not age.isdigit():
        print("Invalid Age. Age should be a numeric value.")
        return False
    return True

def validate_email(email):
    if not "@" in email or not "." in email:
        print("Invalid Email. Email should be in a valid format (e.g., example@example.com).")
        return False
    return True

def display_success_message(roll_number, name, age, email):
    print("\nSuccessfully stored the following data:")
    print("Roll Number:", roll_number)
    print("Name:", name)
    print("Age:", age)
    print("Email:", email)

def read_existing_data(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        return []
    
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

def save_to_excel(file_name, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row_data in data:
        sheet.append(row_data)

    workbook.save(file_name)

if __name__ == "__main__":
    timestamp = datetime.now().strftime("%d%m%y")
    file_name = f"student_data_{timestamp}.xlsx"

    existing_data = read_existing_data(file_name)

    while True:
        roll_number, name, age, email = get_input()

        if not validate_roll_number(roll_number):
            continue

        if not validate_age(age):
            continue

        if not validate_email(email):
            continue

        display_success_message(roll_number, name, age, email)

        # Append the new data to the existing_data
        existing_data.append([roll_number, name, age, email])

        save_to_excel(file_name, existing_data)

        another_entry = input("\nDo you want to enter another data (yes/no)? ").lower()
        if another_entry != "yes":
            print("Exiting program.")
            break
